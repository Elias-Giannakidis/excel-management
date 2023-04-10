# import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import excel.testJson as json

data = json.data

def getTestWs():
    path = 'excel/test.xlsx'
    wb = load_workbook(path)
    # Work sheet
    ws = wb.active
    return ws

def test1():
    path = 'excel/test.xlsx'
    wb = load_workbook(path)

    # Work sheet
    ws = wb.active
    print(ws)
    print('------')
    print(ws['A1'].value)

    #write on excel
    ws['A2'].value = "Test"
    wb.save(path)

    #print all the sheetnames
    print(wb.sheetnames)
    ws = wb['Grade']

    # create a new sheet
    wb.create_sheet("myTest")
    wb.save(path)

def createExcel():
    # Create a new excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    row = [ 'Elias', 'Test', 'column3']
    ws.append(row)
    wb.save('excel/elias.xlsx')

def readExcelValues():
    path = 'excel/test.xlsx'
    wb = load_workbook(path)
    # Work sheet
    ws = wb.active
    for row in range(1, 11):
        for col in range(1, 5):
            char = get_column_letter(col)
            print(ws[char + str(row)].value)

    ws.merge_cells("A1:D1")
    ws.unmerge_cells("A1:D1")

    #insert empty row
    ws.insert_rows(7)
    ws.delete_rows(7)
    ws.insert_cols(2)
    ws.delete_cols(2)

    ws.move_range("C1:D11", rows=2, cols=3)

def styles():
    title = "excel/newExcel.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"
    data = json.data
    headings = ['Name'] + list(data["Joe"].keys())
    print(headings)
    ws.append(headings)
    for person in data:
        grades = list(data[person].values())
        ws.append([person] + grades)

    for col in range(2, len(data['Joe']) + 2):
        char = get_column_letter(col)
        ws[char + "5"] = f"=SUM({char + '2'}:{char + '4'})/{len(data)}"

    for col in range(1, 5):
        ws[get_column_letter(col) + '1'].font = Font(bold=True)
    wb.save(title)


def main():
    styles()
if __name__ == '__main__':
    main()